"""Fill the master Excel template with extracted data using openpyxl (preserves all styles)."""

from __future__ import annotations

import json
import os
from copy import copy
from dataclasses import asdict
from datetime import datetime
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

# openpyxl ≤ 3.1.x doesn't accept 'extLst' that Excel embeds in style XML.
# Patch PatternFill so unknown kwargs are silently ignored.
_original_patternfill_init = PatternFill.__init__


def _patternfill_init_patched(self, *args, **kwargs):
    kwargs.pop("extLst", None)
    _original_patternfill_init(self, *args, **kwargs)


PatternFill.__init__ = _patternfill_init_patched

from src.paths import APP_DIR, BUNDLE_DIR
from src.types import (
    AllExtractedData,
    DLData,
    EstimateData,
    FitnessCertData,
    InsuranceData,
    InvoiceData,
    InvoicePart,
    RCData,
    RoutePermitData,
)

PROJECT_ROOT = APP_DIR
TEMPLATE_PATH = BUNDLE_DIR / "templates" / "TEMPLATE.xlsx"

with open(BUNDLE_DIR / "config" / "cellmap.json", "r", encoding="utf-8") as _f:
    CELLMAP: dict[str, Any] = json.load(_f)


# ─── Helpers ─────────────────────────────────────────────────────────────────


def _date_to_excel_serial(date_str: str) -> int | None:
    """Parse 'DD.MM.YYYY' to an Excel serial date number."""
    if not date_str:
        return None
    parts = date_str.split(".")
    if len(parts) != 3:
        return None
    try:
        dd, mm, yyyy = int(parts[0]), int(parts[1]), int(parts[2])
        dt = datetime(yyyy, mm, dd)
        epoch = datetime(1899, 12, 30)  # Excel epoch with 1900 leap-year bug
        return (dt - epoch).days
    except (ValueError, OverflowError):
        return None


def _write_cell(ws: Worksheet, cell_addr: str, value: Any) -> None:
    """Write a value to a cell, preserving existing formatting.

    Skip formula cells and empty values.
    """
    if value is None or value == "":
        return

    cell = ws[cell_addr]

    # Skip merged sub-cells (only the top-left cell of a merge is writable)
    from openpyxl.cell.cell import MergedCell

    if isinstance(cell, MergedCell):
        return

    # Never overwrite formula cells
    if cell.data_type == "f" or (
        cell.value and isinstance(cell.value, str) and cell.value.startswith("=")
    ):
        return

    # Preserve the existing style
    cell.value = value


def _build_allotment_text(insurer_name: str) -> str:
    text = insurer_name.replace("Co. Ltd.", "Co. Ltd.,").replace("Co Ltd", "Co Ltd,")
    return f"Case allotted by {text} RO: SVC, Sector-17, Chd."


def _to_num(val: Any, default: float = 0.0) -> float | int:
    """Safely convert to number."""
    try:
        n = float(val)
        return int(n) if n == int(n) else n
    except (ValueError, TypeError):
        return default


def _shift_formula_refs(formula: str, min_row: int, offset: int) -> str:
    """Shift all cell row-references >= min_row by offset in an Excel formula."""
    import re as _re

    if offset == 0:
        return formula

    def _replace(m: _re.Match) -> str:
        col = m.group(1)
        row = int(m.group(2))
        if row >= min_row:
            row += offset
        return f"{col}{row}"

    return _re.sub(r"([A-Z]+)(\d+)", _replace, formula)


def _fix_shifted_formulas(ws: Worksheet, from_row: int, offset: int) -> None:
    """After insert_rows, fix every formula from `from_row` to end of sheet."""
    if offset == 0:
        return
    for row in range(from_row, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.data_type == "f" or (
                isinstance(cell.value, str) and str(cell.value).startswith("=")
            ):
                cell.value = _shift_formula_refs(cell.value, from_row - offset, offset)


def _shift_merged_cells(ws: Worksheet, from_row: int, offset: int) -> None:
    """After insert_rows, shift all merged-cell ranges that start at or after
    from_row down by offset.  Ranges that span the insertion point are expanded.
    openpyxl does NOT do this automatically.

    Directly mutates the CellRange coordinate fields rather than calling
    unmerge_cells/merge_cells, because those helpers try to delete/create entries
    in ws._cells — which have already been physically re-keyed by insert_rows and
    would produce a KeyError for any sub-cell that was moved."""
    if offset == 0:
        return
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= from_row:
            mr.min_row += offset
            mr.max_row += offset
        elif mr.max_row >= from_row:
            mr.max_row += offset


def _copy_row_styles(ws: Worksheet, src_row: int, dst_row: int) -> None:
    """Copy cell styles (font, border, fill, alignment, number_format) from
    src_row to dst_row for every column that has styling in the source."""
    for col_idx in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=col_idx)
        dst = ws.cell(row=dst_row, column=col_idx)
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


# ─── Section fillers ─────────────────────────────────────────────────────────


def _fill_insurance(ws: Worksheet, data: InsuranceData) -> None:
    mapping = CELLMAP["sheet1"]["insurance"]
    data_dict = asdict(data)
    for cell, field in mapping.items():
        if field == "insurer_allotment_text":
            _write_cell(
                ws,
                cell,
                _build_allotment_text(data.insurer_name) if data.insurer_name else "",
            )
        elif field == "idv":
            _write_cell(ws, cell, _to_num(data.idv))
        else:
            _write_cell(ws, cell, data_dict.get(field, ""))


def _fill_rc(ws: Worksheet, data: RCData) -> None:
    mapping = CELLMAP["sheet1"]["rc"]
    data_dict = asdict(data)
    for cell, field in mapping.items():
        if field == "fuel_type_label":
            _write_cell(
                ws, cell, f"Fuel used: {data.fuel_type}" if data.fuel_type else ""
            )
        elif field in ("cubic_capacity", "seating_capacity"):
            _write_cell(ws, cell, _to_num(data_dict.get(field, 0)))
        else:
            _write_cell(ws, cell, data_dict.get(field, ""))

    # Hidden serial cell — drives life-of-vehicle formula
    serial_map = CELLMAP["sheet1"].get("rc_date_serial", {})
    for cell, field in serial_map.items():
        if field == "date_of_reg_issue_serial":
            serial = _date_to_excel_serial(data.date_of_reg_issue)
            if serial is not None:
                _write_cell(ws, cell, serial)


def _fill_dl(ws: Worksheet, data: DLData) -> None:
    mapping = CELLMAP["sheet1"]["dl"]
    data_dict = asdict(data)
    for cell, field in mapping.items():
        _write_cell(ws, cell, data_dict.get(field, ""))


def _fill_fitness_cert(ws: Worksheet, data: FitnessCertData) -> None:
    mapping = CELLMAP["sheet1"].get("fitness_cert", {})
    data_dict = asdict(data)
    for cell, field in mapping.items():
        _write_cell(ws, cell, data_dict.get(field, ""))


def _fill_route_permit(ws: Worksheet, data: RoutePermitData) -> None:
    mapping = CELLMAP["sheet1"].get("route_permit", {})
    data_dict = asdict(data)
    for cell, field in mapping.items():
        _write_cell(ws, cell, data_dict.get(field, ""))


def _fill_workshop(
    ws: Worksheet,
    estimate: Optional[EstimateData],
    invoice: Optional[InvoiceData],
) -> None:
    """Fill workshop / place of survey (G63, G64) from estimate or invoice."""
    mapping = CELLMAP["sheet1"].get("workshop", {})
    if not mapping:
        return
    # Prefer estimate data; fall back to invoice
    dealer_name = ""
    dealer_address = ""
    if estimate:
        dealer_name = estimate.dealer_name or ""
        dealer_address = estimate.dealer_address or ""
    if not dealer_name and invoice:
        dealer_name = invoice.dealer_name or ""
    if not dealer_address and invoice:
        dealer_address = invoice.dealer_address or ""

    for cell, field in mapping.items():
        if field == "dealer_name":
            _write_cell(ws, cell, dealer_name)
        elif field == "dealer_address":
            _write_cell(ws, cell, dealer_address)


def _match_parts_ai(
    estimate_parts: list,
    invoice_parts: list[InvoicePart],
) -> dict[int, int]:
    """Use AI to match estimate parts to invoice parts by name.

    Returns {estimate_index: invoice_index} mapping.
    Parts that don't appear in the invoice are not in the mapping.
    """
    from src.utils.ai_client import vision_extract_json

    if not estimate_parts or not invoice_parts:
        return {}

    est_list = [{"idx": i, "name": p.name} for i, p in enumerate(estimate_parts)]
    inv_list = [{"idx": j, "name": p.name} for j, p in enumerate(invoice_parts)]

    prompt = f"""Match vehicle parts from an Estimate list to an Invoice list.
These are auto parts from Indian automobile repair documents.
Parts may have slightly different names but refer to the same physical part.
Only match parts that are genuinely the same part. Do NOT force matches for unrelated parts.

Estimate parts:
{json.dumps(est_list)}

Invoice parts:
{json.dumps(inv_list)}

Return a JSON object with a single key "matches" containing an array of objects.
Each object has "estimate_idx" and "invoice_idx" for parts that match.
Only include matches where you are confident the parts are the same.
Parts in the estimate that do NOT appear in the invoice should be omitted.

Example: {{"matches": [{{"estimate_idx": 0, "invoice_idx": 2}}, {{"estimate_idx": 1, "invoice_idx": 0}}]}}"""

    try:
        data = vision_extract_json([], prompt, max_output_tokens=65536)
        matches = data.get("matches", [])
        result: dict[int, int] = {}
        used_inv: set[int] = set()
        for m in matches:
            ei = m.get("estimate_idx")
            ii = m.get("invoice_idx")
            if (
                isinstance(ei, int)
                and isinstance(ii, int)
                and 0 <= ei < len(estimate_parts)
                and 0 <= ii < len(invoice_parts)
                and ii not in used_inv
            ):
                result[ei] = ii
                used_inv.add(ii)
        return result
    except Exception as e:  # pylint: disable=broad-except
        print(f"    ⚠ AI part matching failed: {e}, all invoice prices set to N.A.")
        return {}


def _fill_parts_table(  # pylint: disable=too-many-locals
    ws: Worksheet, estimate: EstimateData, invoice: Optional[InvoiceData]
) -> int:
    """Fill parts table. Returns the number of extra rows inserted."""
    cfg = CELLMAP["sheet1"]["parts_table"]
    start_row: int = cfg["start_row"]
    max_slots: int = cfg.get("max_slots", 6)
    parts = estimate.parts or []

    # AI-based matching FIRST so we know how many unmatched invoice items exist
    inv_parts: list[InvoicePart] = (
        list(invoice.parts_assessed) if invoice and invoice.parts_assessed else []
    )
    part_match: dict[int, int] = {}
    if inv_parts:
        print("    Matching estimate parts → invoice parts (AI)...")
        part_match = _match_parts_ai(parts, inv_parts)
        print(f"    Matched {len(part_match)} of {len(parts)} estimate parts")

    # Count unmatched invoice items that will be appended
    matched_inv_indices = set(part_match.values())
    unmatched_inv_count = sum(
        1 for j in range(len(inv_parts)) if j not in matched_inv_indices
    )

    # Insert extra rows if TOTAL items exceed available slots
    total_items = len(parts) + unmatched_inv_count
    extra = max(0, total_items - max_slots)
    if extra > 0:
        # Insert WITHIN the data range so SUM formulas auto-expand.
        # Inserting before the last slot (start_row + max_slots - 1) ensures
        # the range end reference shifts, expanding =SUM(F98:F103) correctly.
        insert_at = start_row + max_slots - 1
        ws.insert_rows(insert_at, extra)
        # openpyxl insert_rows does NOT update formula strings — fix them.
        _fix_shifted_formulas(ws, insert_at + extra, extra)
        # openpyxl insert_rows does NOT update merged-cell ranges — fix them.
        # Use insert_at (not insert_at + extra) as from_row: merges are still at
        # their original positions (openpyxl never auto-shifts them), so we must
        # shift every merge at or below the insertion point.
        _shift_merged_cells(ws, insert_at, extra)
        # openpyxl insert_rows creates blank rows — copy styles from template row.
        for r in range(insert_at, insert_at + extra):
            _copy_row_styles(ws, start_row, r)

    # Write column headers in the row above the data
    header_row = start_row - 1
    _write_cell(ws, f"J{header_row}", "Extra")
    _write_cell(ws, f"K{header_row}", "Invoice No")
    ws[f"K{header_row}"].alignment = Alignment(horizontal="right")

    for i, part in enumerate(parts):
        row = start_row + i
        sn = part.sn if part.sn else i + 1

        _write_cell(ws, f"B{row}", sn)
        _write_cell(ws, f"C{row}", part.name)
        _write_cell(ws, f"F{row}", _to_num(part.estimated_price))

        # Ensure the part name cell (C) inherits styling from the styled B cell
        b_cell = ws[f"B{row}"]
        c_cell = ws[f"C{row}"]
        c_cell.font = copy(b_cell.font)
        c_cell.border = copy(b_cell.border)
        c_cell.alignment = copy(b_cell.alignment)

        # Look up AI match
        assessed_price: float | None = None
        if i in part_match:
            assessed_price = _to_num(inv_parts[part_match[i]].assessed_price)
            # Write the invoice serial number (1-based)
            _write_cell(ws, f"K{row}", part_match[i] + 1)
            ws[f"K{row}"].alignment = Alignment(horizontal="right")

        cat = (part.category or "").lower()
        est_num = _to_num(part.estimated_price)

        if assessed_price is not None:
            # Cap assessed at estimated; excess goes to column J
            if est_num and est_num > 0:
                capped = min(assessed_price, est_num)
                excess = max(0, assessed_price - est_num)
            else:
                capped = assessed_price
                excess = 0

            if cat == "metal":
                _write_cell(ws, f"G{row}", capped)
            elif cat == "glass":
                _write_cell(ws, f"I{row}", capped)
            else:
                _write_cell(ws, f"H{row}", capped)

            if excess > 0:
                _write_cell(ws, f"J{row}", excess)
        else:
            if cat == "metal":
                _write_cell(ws, f"G{row}", "N.A.")
            elif cat == "glass":
                _write_cell(ws, f"I{row}", "N.A.")
            else:
                _write_cell(ws, f"H{row}", "N.A.")

    # Append unmatched invoice items (invoice-only) after estimate parts
    next_row = start_row + len(parts)
    next_sn = len(parts) + 1
    for j, inv_part in enumerate(inv_parts):
        if j not in matched_inv_indices:
            row = next_row
            _write_cell(ws, f"B{row}", next_sn)
            _write_cell(ws, f"C{row}", inv_part.name)
            # Ensure the part name cell (C) inherits styling from the styled B cell
            b_cell = ws[f"B{row}"]
            c_cell = ws[f"C{row}"]
            c_cell.font = copy(b_cell.font)
            c_cell.border = copy(b_cell.border)
            c_cell.alignment = copy(b_cell.alignment)
            # Write invoice serial number (1-based)
            _write_cell(ws, f"K{row}", j + 1)
            ws[f"K{row}"].alignment = Alignment(horizontal="right")
            # Full amount goes to extra (col J) — no estimate to cap against
            inv_price = _to_num(inv_part.assessed_price)
            if inv_price:
                _write_cell(ws, f"J{row}", inv_price)
            next_row += 1
            next_sn += 1

    # Add SUM formula for the Extra column (J) subtotal
    last_data_row = start_row + max_slots - 1 + extra  # template last slot, shifted
    subtotal_row = last_data_row + 1  # row 135 in base template
    ws[f"J{subtotal_row}"] = f"=SUM(J{start_row}:J{last_data_row})"

    return extra


def _offset_cell(cell_addr: str, row_offset: int) -> str:
    """Shift a cell address like 'I117' down by row_offset rows."""
    if row_offset == 0:
        return cell_addr
    import re as _re

    m = _re.match(r"([A-Z]+)(\d+)", cell_addr)
    if not m:
        return cell_addr
    return f"{m.group(1)}{int(m.group(2)) + row_offset}"


def _fill_labour_table(
    ws: Worksheet,
    estimate: EstimateData,
    row_offset: int = 0,
) -> None:
    cfg = CELLMAP["sheet1"]["labour_table"]
    start_row: int = cfg["start_row"] + row_offset
    labour = estimate.labour or []

    for i, item in enumerate(labour):
        row = start_row + i
        _write_cell(ws, f"C{row}", item.description)

        # Ensure the description cell (C) inherits styling from the styled B cell
        b_cell = ws[f"B{row}"]
        c_cell = ws[f"C{row}"]
        c_cell.font = copy(b_cell.font)
        c_cell.border = copy(b_cell.border)
        c_cell.alignment = copy(b_cell.alignment)

        # Only write non-zero labour values (leave cell empty for 0)
        if _to_num(item.rr):
            _write_cell(ws, f"F{row}", _to_num(item.rr))
        if _to_num(item.denting):
            _write_cell(ws, f"G{row}", _to_num(item.denting))
        if _to_num(item.cw):
            _write_cell(ws, f"H{row}", _to_num(item.cw))
        if _to_num(item.painting):
            _write_cell(ws, f"I{row}", _to_num(item.painting))

    if estimate.total_labour_estimated:
        est_cell = cfg.get("total_estimated_cell")
        if est_cell:
            _write_cell(
                ws,
                _offset_cell(est_cell, row_offset),
                _to_num(estimate.total_labour_estimated),
            )


# ─── Main export ─────────────────────────────────────────────────────────────


def fill_excel(
    all_data: AllExtractedData, output_path: str, ref_number: str | None = None
) -> None:
    """Load master .xlsx template, fill data, save to output_path. All styles preserved."""
    wb = load_workbook(str(TEMPLATE_PATH))

    ws1 = wb["Sheet1"]

    if all_data.insurance:
        _fill_insurance(ws1, all_data.insurance)

    if all_data.rc:
        _fill_rc(ws1, all_data.rc)

    if all_data.dl:
        _fill_dl(ws1, all_data.dl)

    if all_data.fitness_cert:
        _fill_fitness_cert(ws1, all_data.fitness_cert)

    if all_data.route_permit:
        _fill_route_permit(ws1, all_data.route_permit)

    # Fill workshop / place of survey from estimate or invoice
    _fill_workshop(ws1, all_data.estimate, all_data.invoice)

    # Fill estimate AFTER insurance/rc/dl — parts insertion shifts rows below,
    # but all insurance/rc/dl cells are above the insertion point.
    row_offset = 0
    if all_data.estimate:
        row_offset = _fill_parts_table(ws1, all_data.estimate, all_data.invoice)
        _fill_labour_table(ws1, all_data.estimate, row_offset)

    if ref_number:
        ref = f"SK/2025-26/OICL/{ref_number}"
        _write_cell(ws1, "C8", ref)

    # Other sheets (Sheet2, Sheet3, Sheet4, Sheet5, Sheet7) use formulas
    # referencing Sheet1, so they auto-populate — no manual fill needed.

    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    print(f"  ✓ Excel saved: {output_path}")

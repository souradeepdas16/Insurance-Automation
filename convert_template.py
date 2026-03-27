"""One-time script: convert MASTER_OIC_TATA.xls (legacy) to .xlsx preserving data and structure.

Run once:  python convert_template.py

Note: xlrd can read .xls but does NOT transfer styles to openpyxl.
      For perfect style preservation, open the .xls in Excel/LibreOffice and
      Save As .xlsx manually. This script transfers cell values, merged cells,
      column widths, and row heights — which is the best possible without Excel COM.
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import xlrd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

SRC = Path(__file__).parent / "templates" / "MASTER_OIC_TATA.xls"
DST = Path(__file__).parent / "templates" / "MASTER_OIC_TATA.xlsx"


def convert() -> None:  # pylint: disable=too-many-locals,too-many-branches
    if not SRC.exists():
        print(f"Source not found: {SRC}")
        sys.exit(1)

    xls = xlrd.open_workbook(str(SRC), formatting_info=True)
    wb = Workbook()

    for idx, sheet_name in enumerate(xls.sheet_names()):
        ws_src = xls.sheet_by_index(idx)

        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Copy cell values
        for row_idx in range(ws_src.nrows):
            for col_idx in range(ws_src.ncols):
                cell = ws_src.cell(row_idx, col_idx)
                value = cell.value

                # xlrd cell types: 0=empty, 1=text, 2=number, 3=date, 4=boolean, 5=error, 6=blank
                if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    continue

                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        date_tuple = xlrd.xldate_as_tuple(value, xls.datemode)
                        value = datetime(*date_tuple[:6])
                    except Exception:  # pylint: disable=broad-except
                        pass

                ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)

        # Copy merged cell ranges
        for crange in ws_src.merged_cells:
            r1, r2, c1, c2 = crange
            ws.merge_cells(
                start_row=r1 + 1,
                start_column=c1 + 1,
                end_row=r2,
                end_column=c2,
            )

        # Copy column widths (approximate)
        for col_idx in range(ws_src.ncols):
            col_letter = get_column_letter(col_idx + 1)
            try:
                width = ws_src.colinfo_map.get(col_idx)
                if width:
                    ws.column_dimensions[col_letter].width = width.width / 256
            except Exception:  # pylint: disable=broad-except
                pass

        # Copy row heights
        for row_idx in range(ws_src.nrows):
            try:
                row_info = ws_src.rowinfo_map.get(row_idx)
                if row_info and row_info.height:
                    ws.row_dimensions[row_idx + 1].height = row_info.height / 20
            except Exception:  # pylint: disable=broad-except
                pass

    wb.save(str(DST))
    print(f"✓ Converted: {DST}")
    print()
    print("IMPORTANT: For perfect style/formatting preservation, open this file")
    print("in Excel or LibreOffice, verify it looks correct, then save it again.")
    print("The programmatic conversion preserves data, merges, and dimensions,")
    print("but fonts/borders/colors require manual save-as from Excel.")


if __name__ == "__main__":
    convert()

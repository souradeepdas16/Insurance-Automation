"""Create an empty template from sample_excel.xlsx by clearing data cells."""

import openpyxl
from openpyxl.styles import PatternFill

_orig = PatternFill.__init__


def _patched(self, *a, **kw):
    kw.pop("extLst", None)
    _orig(self, *a, **kw)


PatternFill.__init__ = _patched

from openpyxl.cell.cell import MergedCell

wb = openpyxl.load_workbook("templates/sample_excel.xlsx")
ws = wb["Sheet1"]


def clear_cell(ws, addr):
    """Clear a cell's value but keep style. Skip formulas and merged sub-cells."""
    cell = ws[addr]
    if isinstance(cell, MergedCell):
        return
    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
        return  # Keep formulas
    cell.value = None


# ── Insurance data cells ──
for c in ["G13", "G14", "G15", "G16", "K16", "K13", "G18", "G19", "G20", "G22"]:
    clear_cell(ws, c)

# Extra sample data in auxiliary columns
for c in ["N16", "W14", "X14", "Y14", "W15"]:
    clear_cell(ws, c)

# ── RC data cells ──
for c in [
    "G25",
    "G26",
    "G27",
    "G29",
    "G30",
    "G31",
    "G32",
    "G33",
    "G34",  # Pre-accident condition
    "G35",
    "G36",
    "G37",  # Fitness cert
    "G38",  # Permit no
    "G39",  # Valid upto
    "G40",  # Type of permit
    "G41",  # Route
    "G42",  # Seating capacity
    "G43",
    "J43",  # Road tax, fuel type
    "K44",  # Colour
    "G45",  # Cubic capacity
]:
    clear_cell(ws, c)

# ── DL data cells ──
for c in ["G47", "K47", "G48", "G49", "G50", "G51", "G52", "G53", "G54"]:
    clear_cell(ws, c)

# ── Accident data cells ──
for c in ["G57", "N57", "G58"]:
    clear_cell(ws, c)

# ── Survey detail cells ──
for c in ["G61", "G62", "G63", "G64"]:
    clear_cell(ws, c)

# ── Workshop cells ──
for c in ["G68", "G69", "G70"]:
    clear_cell(ws, c)

# ── FIR, Injury ──
for c in ["G72", "G74"]:
    clear_cell(ws, c)

# ── Narrative sections (case-specific text) ──
for c in ["C78", "C81", "C84", "C87"]:
    clear_cell(ws, c)

# ── Parts table data (rows 100-134) ──
for row in range(100, 135):
    for col in ["B", "C", "F", "G", "H", "I", "J"]:
        clear_cell(ws, f"{col}{row}")

# ── Labour data (rows 156-160) ──
for row in range(156, 161):
    for col in ["B", "C", "F", "G", "H", "I"]:
        clear_cell(ws, f"{col}{row}")

# ── Labour estimated total ──
clear_cell(ws, "I154")

# ── Labour adjustments ──
clear_cell(ws, "I162")

# ── Salvage value ──
clear_cell(ws, "G180")

# ── Date cells for vehicle age calculation ──
clear_cell(ws, "E258")
clear_cell(ws, "E259")

# ── Notes section (sample-specific) ──
for c in ["M148", "M149"]:
    clear_cell(ws, c)

# ── Ref number (template default) ──
ws["C8"].value = None  # Will be set per case

# ── Sheet7 (Re-inspection): clear case-specific data ──
ws7 = wb["Sheet7"]
# Clear insurer address in letter
for c in ["B14", "B15"]:
    clear_cell(ws7, c)
# Clear narrative text
clear_cell(ws7, "B22")
# Clear reinspection parts table (rows 30-57)
for row in range(30, 58):
    for col in ["B", "C", "F", "G"]:
        clear_cell(ws7, f"{col}{row}")
# Clear salvage/labour remarks
for c in ["C60", "C63", "C64", "C76"]:
    clear_cell(ws7, c)

# ── Sheet3: clear invoice-specific data ──
ws3 = wb["Sheet3"]
for c in ["G39", "G40", "G41", "G42"]:
    clear_cell(ws3, c)

# Save the empty template
output = "templates/TEMPLATE.xlsx"
wb.save(output)
print(f"Empty template saved to: {output}")
print(f"Sheets: {wb.sheetnames}")
